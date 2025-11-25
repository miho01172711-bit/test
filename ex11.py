import  openpyxl as  op
from  openpyxl.styles.fonts import  Font

wb = op.Workbook()
ws = wb.active

ws["A1"].value = "Font test1"
ws["A1"].font = Font(size=20, italic = True, bold = True)

ws["A2"].value = "Font Test2"
font_format = Font(size=12, name='굴림', color = '000000')
ws["A2"].font = font_format

wb.save("test_result.xlsx")
wb.close()
