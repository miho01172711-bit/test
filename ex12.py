import openpyxl as op
from openpyxl.styles import Alignment, Border, Side

wb = op.Workbook()
ws = wb.active

header = ["이름", "국어", "영어", "수학"]
data = [
    ["홍길동", 50, 80, 60],
    ["강감찬", 80, 70, 60],
    ["김철수", 40, 50, 70],
]

for col, value in enumerate(header, start=1):
    ws.cell(row=1, column=col, value=value)
for row_idx, row_data in enumerate(data, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        ws.cell(row=row_idx, column=col_idx, value=value)

for col, value in enumerate(header, start=1):
    ws.cell(row=1, column=col, value=value)
for row_idx, row_data in enumerate(data, start=2):
    for col_idx, value in enumerate(row_data, start=1):
        ws.cell(row=row_idx, column=col_idx, value=value)

max_row = ws.max_row
ws.cell(row=1, column=max_row+1, value="평균")

for row in range(2, max_row+1): 
    ws.cell(row=row, column=max_row+1, value=f"=AVERAGE(B{row}:D{row})")

thin = Side(border_style="thin", color="000000")
border = Border(left=thin, right=thin, top=thin, bottom=thin)

row_max = ws.max_row
col_max = ws.max_column

for r in range(1, row_max+1): 
    for c in range(1, col_max+1): 
        cell = ws.cell(row=r, column=c)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

wb.save("result.xlsx")
wb.close()