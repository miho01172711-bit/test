import openpyxl as op
wb = op.load_workbook("result.xlsx", data_only=True)
ws = wb.active

max_row = ws.max_row
for row_index in range(2, max_row+1):
    average = ws.cell(row = row_index, column=5).value
    if average >= 70:
        ws.cell(row=row_index, column=6).value = "합격"
    else:
        ws.cell(row=row_index, column=6).value = "불합격"  

from  openpyxl.styles.fonts import  Font
pass_format =  Font(size=12, name='굴림', color = '000000FF')
fail_format =  Font(size=12, name='굴림', color = '00FF0000')

max_row = ws.max_row
for row_index in range(2, max_row+1):
    result_str = ws.cell(row = row_index, column=6).value
    if result_str == "합격":
        ws.cell(row=row_index, column=6).font = pass_format
    else:
        ws.cell(row=row_index, column=6).font = fail_format

wb.save("con_result.xlsx")