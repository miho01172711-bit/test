import openpyxl as op
wb = op.Workbook()
print(wb)
ws = wb.active
ws.title = "업"
wb.create_sheet("무")
wb.create_sheet("자")
wb.create_sheet("동")
wb.create_sheet("화")
wb.save("test.xlsx")
