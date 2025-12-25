import openpyxl as op
wb = op.load_workbook("test.xlsx")
ws = wb["무"]

ws.cell(row=1, column=2).value = "입력테스트1"
ws["C1"] = "입력테스트2"
wb.save("test.xlsx")

datalist = [2,4,8,16,32,64,128,256]
i=1
for data in datalist:
    ws.cell(row=i, column=1).value = data
    i=i+1
wb.save("test.xlsx")