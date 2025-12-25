import openpyxl as op
wb = op.load_workbook("test.xlsx")
ws = wb.active
data1 = ws.cell(row=1, column=2).value
data2 = ws["B1"].value
print("Cell(1,2):", data1)
print("Range(B1):", data2)

rng = ws["A1:B1"]
print(rng)

rng = ws["A1:C3"]
for rng_data in rng:
    for cell_data in rng_data:
        print(cell_data.value)